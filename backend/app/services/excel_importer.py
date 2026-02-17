"""
Excel Rules Importer — reads acquisition-rules-config.xlsx and populates
all rule tables (IntakePaths, Thresholds, ApprovalTemplates, DocumentRules,
AdvisoryTriggers).

The Excel workbook is the single source of truth. When rules change, an admin
re-imports the workbook and the system picks up the changes at runtime.
"""

import json
import os
import openpyxl
from app.extensions import db
from app.models.intake_path import IntakePath
from app.models.advisory_trigger import AdvisoryTriggerRule
from app.models.threshold import ThresholdConfig
from app.models.approval import ApprovalTemplate, ApprovalTemplateStep
from app.models.document import DocumentTemplate, DocumentRule


# ---------------------------------------------------------------------------
# Value normalization maps (Excel human-readable → DB keys)
# ---------------------------------------------------------------------------
ACQ_TYPE_MAP = {
    'new competitive': 'new_competitive',
    'new competitive (urgency)': 'new_competitive_urgency',
    're-compete': 'recompete',
    'follow-on sole source': 'follow_on_sole_source',
    'follow-on ss': 'follow_on_sole_source',
    'sole source': 'sole_source',
    'brand name': 'brand_name',
    'option exercise': 'option_exercise',
    'bridge extension': 'bridge_extension',
    'bilateral modification': 'bilateral_mod',
    'unilateral modification': 'unilateral_mod',
    'clin reallocation': 'clin_reallocation',
    'clin execution — odc': 'clin_execution_odc',
    'clin execution — travel': 'clin_execution_travel',
    'clin execution + funding action': 'clin_execution_funding',
    'clin exec + funding': 'clin_execution_funding',
}

PIPELINE_MAP = {
    'full': 'full',
    'full + legal': 'full_legal',
    'abbreviated': 'abbreviated',
    'ko-abbreviated': 'ko_abbreviated',
    'ko-only': 'ko_only',
    'clin-execution': 'clin_execution',
    'clin-exec-funding': 'clin_exec_funding',
    'depends on $ change': 'depends_on_value',
}

TIER_MAP = {
    'micro-purchase': 'micro',
    'micro': 'micro',
    'sat': 'sat',
    'above sat': 'above_sat',
    'major': 'major',
    'all': None,  # wildcard
}

BUY_CATEGORY_MAP = {
    'product': 'product',
    'service': 'service',
    'software/license': 'software_license',
    'software': 'software_license',
    'mixed': 'mixed',
    'all': None,
}

CONTRACT_CHARACTER_MAP = {
    'product': 'product',
    'service': 'service',
    'mixed-product': 'mixed_product',
    'mixed-service': 'mixed_service',
    'all': None,
}

NEED_TYPE_MAP = {
    'new': 'new',
    'continue/extend': 'continue_extend',
    'change existing': 'change_existing',
}


def _normalize_acq_type(label):
    """Convert an Excel acq type label to a DB key."""
    if not label:
        return None
    return ACQ_TYPE_MAP.get(label.strip().lower(), label.strip().lower().replace(' ', '_').replace('—', '_'))


def _normalize_pipeline(label):
    """Convert an Excel pipeline label to a DB key."""
    if not label:
        return None
    return PIPELINE_MAP.get(label.strip().lower(), label.strip().lower().replace(' ', '_').replace('-', '_'))


def _normalize_tier(label):
    """Convert an Excel tier label to a DB key."""
    if not label:
        return None
    return TIER_MAP.get(label.strip().lower(), label.strip().lower().replace(' ', '_'))


def _normalize_buy_category(label):
    """Convert an Excel buy category label to a DB key."""
    if not label:
        return None
    return BUY_CATEGORY_MAP.get(label.strip().lower(), label.strip().lower().replace('/', '_'))


def _normalize_need_type(label):
    """Convert an Excel need type label to a DB key."""
    if not label:
        return None
    return NEED_TYPE_MAP.get(label.strip().lower(), label.strip().lower().replace(' ', '_').replace('/', '_'))


def _parse_comma_list(text):
    """Split a comma-separated string into a list, stripping whitespace."""
    if not text or text.strip() in ('-', '', 'None', 'ALL'):
        return []
    return [item.strip() for item in str(text).split(',') if item.strip()]


def _normalize_acq_type_list(text):
    """Parse a comma-separated list of acq types from the Excel and normalize each."""
    if not text or text.strip() in ('-', '', 'None'):
        return []
    # Handle "ALL" and "ALL (not X, not Y)" patterns
    raw = str(text).strip()
    if raw.upper().startswith('ALL'):
        return ['__ALL__']
    items = [item.strip() for item in raw.split(',') if item.strip()]
    return [_normalize_acq_type(item) for item in items if _normalize_acq_type(item)]


def _normalize_tier_list(text):
    """Parse a comma-separated list of tiers and normalize each."""
    if not text or text.strip() in ('-', '', 'None'):
        return []
    raw = str(text).strip()
    if raw.upper() == 'ALL':
        return ['__ALL__']
    items = [item.strip() for item in raw.split(',') if item.strip()]
    return [_normalize_tier(item) for item in items if _normalize_tier(item)]


def _normalize_buy_category_list(text):
    """Parse a comma-separated list of buy categories and normalize each."""
    if not text or text.strip() in ('-', '', 'None'):
        return []
    raw = str(text).strip()
    if raw.upper() == 'ALL':
        return ['__ALL__']
    items = [item.strip() for item in raw.split(',') if item.strip()]
    return [_normalize_buy_category(item) for item in items if _normalize_buy_category(item)]


def _normalize_character_list(text):
    """Parse a comma-separated list of contract characters and normalize each."""
    if not text or text.strip() in ('-', '', 'None'):
        return []
    raw = str(text).strip()
    if raw.upper() == 'ALL':
        return ['__ALL__']
    items = [item.strip() for item in raw.split(',') if item.strip()]
    result = []
    for item in items:
        key = CONTRACT_CHARACTER_MAP.get(item.lower(), item.lower().replace('-', '_').replace(' ', '_'))
        if key:
            result.append(key)
    return result


# ---------------------------------------------------------------------------
# Sheet importers
# ---------------------------------------------------------------------------

def import_intake_paths(wb):
    """Import Intake Paths sheet → IntakePath table."""
    ws = wb['Intake Paths']
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows (3)
        path_id = row[0]
        if not path_id or not str(path_id).startswith('PATH-'):
            continue
        path = IntakePath(
            path_id=str(path_id),
            q1_need_type=_normalize_need_type(row[1]),
            q2_situation=str(row[2]).strip() if row[2] else None,
            q3_specific_vendor=str(row[3]).strip() if row[3] and str(row[3]).strip() != '-' else None,
            buy_category=_normalize_buy_category(row[4]) if row[4] and str(row[4]).strip() != '-' else None,
            derived_acq_type=_normalize_acq_type(row[5]) or str(row[5] or '').strip(),
            derived_pipeline=_normalize_pipeline(row[6]) or str(row[6] or '').strip(),
            doc_template_set=str(row[7]).strip() if row[7] else None,
            approval_template_key=str(row[8]).strip() if row[8] else None,
            advisory_triggers=str(row[9]).strip() if row[9] and str(row[9]).strip() != 'None' else None,
            notes=str(row[10]).strip() if row[10] else None,
        )
        db.session.add(path)
        count += 1
    db.session.flush()
    print(f'  Imported {count} intake paths')
    return count


def import_thresholds(wb):
    """Import Thresholds sheet → ThresholdConfig table."""
    ws = wb['Thresholds']
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows (3)
        name = row[0]
        if not name or name == 'Threshold Name':
            continue
        # Normalize name to snake_case DB key
        db_name = str(name).strip().lower().replace(' ', '_').replace('(', '').replace(')', '')
        # Clean up common patterns
        db_name = db_name.replace('micro-purchase_threshold', 'micro_purchase')
        db_name = db_name.replace('simplified_acquisition_threshold_sat', 'simplified_acquisition')
        db_name = db_name.replace('above_sat_ceiling', 'above_sat')
        db_name = db_name.replace('j&a_approval_—_ko_authority', 'ja_ko_authority')
        db_name = db_name.replace('j&a_approval_—_competition_advocate', 'ja_competition_advocate')
        db_name = db_name.replace('j&a_approval_—_head_of_activity', 'ja_head_of_activity')
        db_name = db_name.replace('j&a_approval_—_spe/agency_head', 'ja_spe_agency_head')
        db_name = db_name.replace('investment_threshold_o&m_vs._investment', 'investment_threshold')
        db_name = db_name.replace('scls_threshold', 'scls_threshold')
        db_name = db_name.replace('clin_execution_—_auto-approve_limit', 'clin_auto_approve')

        dollar_val = float(row[1]) if row[1] else 0

        t = ThresholdConfig(
            name=db_name,
            dollar_limit=dollar_val,
            effective_date=str(row[2]).strip() if row[2] else None,
            end_date=str(row[3]).strip() if row[3] else None,
            far_reference=str(row[4]).strip() if row[4] else None,
            tier_determines=str(row[5]).strip() if row[5] else None,
            description=str(row[6]).strip() if row[6] else str(name).strip(),
        )
        db.session.add(t)
        count += 1
    db.session.flush()
    print(f'  Imported {count} thresholds')
    return count


def import_approval_templates(wb):
    """Import Approval Templates sheet → ApprovalTemplate + ApprovalTemplateStep tables."""
    ws = wb['Approval Templates']
    templates = {}  # template_key → ApprovalTemplate instance
    step_count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows (3)
        template_id = row[0]
        if not template_id or not str(template_id).startswith('APPR-'):
            continue

        template_key = str(template_id).strip()
        template_name = str(row[1]).strip() if row[1] else template_key

        # Create template if we haven't seen this key yet
        if template_key not in templates:
            # Derive pipeline_type from template_key
            pipeline_map = {
                'APPR-FULL': 'full',
                'APPR-FULL-LEGAL': 'full_legal',
                'APPR-OPTION': 'abbreviated',
                'APPR-BRIDGE': 'ko_abbreviated',
                'APPR-KO-ONLY': 'ko_only',
                'APPR-CLIN-EXEC': 'clin_execution',
                'APPR-CLIN-EXEC-FUND': 'clin_exec_funding',
                'APPR-MOD': 'modification',
                'APPR-MICRO': 'micro',
            }
            tmpl = ApprovalTemplate(
                template_key=template_key,
                name=template_name,
                pipeline_type=pipeline_map.get(template_key, template_key.lower()),
            )
            db.session.add(tmpl)
            db.session.flush()
            templates[template_key] = tmpl

        tmpl = templates[template_key]
        step_order = int(row[2]) if row[2] else 1
        step_name = str(row[3]).strip() if row[3] else ''
        approver_role = str(row[4]).strip() if row[4] else ''

        # Map human-readable roles to DB role keys
        role_map = {
            'Branch Chief': 'branch_chief',
            'Acquisition Review Board': 'branch_chief',
            'Budget Officer': 'budget',
            'Contracting Officer': 'ko',
            'CIO': 'cto',
            'Component Head': 'branch_chief',
            'General Counsel': 'legal',
            'COR': 'branch_chief',
            'Program Manager': 'branch_chief',
            'CTO': 'cto',
            'Financial Manager': 'budget',
            'Business Manager': 'budget',
            'Supervisor': 'branch_chief',
            'GPC Holder': 'budget',
        }
        approver_role_key = role_map.get(approver_role, approver_role.lower().replace(' ', '_'))

        is_required = str(row[5]).strip().lower() == 'yes' if row[5] else True
        condition_text = str(row[6]).strip() if row[6] else None
        is_conditional = (not is_required) or (condition_text and condition_text.lower() not in ('always', 'none', ''))
        sla_days = int(row[7]) if row[7] else 5
        escalation_to = str(row[8]).strip() if row[8] else None

        # Convert condition text to JSON rule
        condition_rule = None
        if is_conditional and condition_text and condition_text.lower() not in ('always', ''):
            if 'it acquisition' in condition_text.lower():
                condition_rule = json.dumps({'allOf': [
                    {'field': 'intake_q_buy_category', 'operator': 'in',
                     'values': ['software_license', 'service', 'mixed', 'product']}
                ]})
            elif 'major' in condition_text.lower() and '>$9m' in condition_text.lower().replace(' ', ''):
                condition_rule = json.dumps({'allOf': [
                    {'field': 'derived_tier', 'operator': '==', 'value': 'major'}
                ]})
            elif 'major' in condition_text.lower():
                condition_rule = json.dumps({'allOf': [
                    {'field': 'derived_tier', 'operator': '==', 'value': 'major'}
                ]})

        step = ApprovalTemplateStep(
            template_id=tmpl.id,
            step_number=step_order,
            step_name=step_name,
            approver_role=approver_role_key,
            sla_days=sla_days,
            is_conditional=bool(is_conditional and condition_rule),
            condition_rule=condition_rule,
            escalation_to=escalation_to,
        )
        db.session.add(step)
        step_count += 1

    db.session.flush()
    print(f'  Imported {len(templates)} approval templates with {step_count} steps')
    return len(templates), step_count


def import_document_rules(wb):
    """Import Document Rules sheet → DocumentTemplate + DocumentRule tables.

    Converts the Excel column-based conditions into the existing JSON allOf format.
    """
    ws = wb['Document Rules']
    templates_created = 0
    rules_created = 0

    # Gate name normalization
    gate_map = {
        'ASR': 'asr',
        'ISS': 'iss',
        'KO Review': 'ko_review',
        'KO Action': 'ko_review',
        'Finance Gate': 'finance',
        'PM Approval': 'iss',
        'COR Validation': 'ko_review',
    }

    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows (3)
        rule_id = row[0]
        if not rule_id or not str(rule_id).startswith('DOC-'):
            continue

        doc_name = str(row[1]).strip() if row[1] else ''
        doc_category = str(row[2]).strip().lower().replace('-', '_') if row[2] else 'pre_award'
        acq_types_raw = str(row[3]).strip() if row[3] else ''
        tiers_raw = str(row[4]).strip() if row[4] else ''
        buy_cats_raw = str(row[5]).strip() if row[5] else ''
        character_raw = str(row[6]).strip() if row[6] else ''
        applicability = str(row[7]).strip().lower() if row[7] else 'required'
        gate_raw = str(row[8]).strip() if row[8] else ''
        ai_assistable = str(row[9]).strip().lower() in ('yes', 'true') if row[9] else False

        # Create a doc_type_key from the rule_id + name
        doc_type_key = str(rule_id).lower().replace('-', '_')

        gate = gate_map.get(gate_raw, gate_raw.lower().replace(' ', '_')) if gate_raw else 'asr'

        # Create or find DocumentTemplate
        existing = DocumentTemplate.query.filter_by(doc_type_key=doc_type_key).first()
        if not existing:
            tmpl = DocumentTemplate(
                doc_type_key=doc_type_key,
                name=doc_name,
                category=doc_category,
                required_before_gate=gate,
                sort_order=templates_created + 1,
                ai_assistable=ai_assistable,
            )
            db.session.add(tmpl)
            db.session.flush()
            templates_created += 1
        else:
            tmpl = existing

        # Build JSON conditions from the Excel columns
        conditions = []

        # Acq type condition
        acq_types = _normalize_acq_type_list(acq_types_raw)
        if acq_types and '__ALL__' not in acq_types:
            conditions.append({
                'field': 'derived_acquisition_type',
                'operator': 'in',
                'values': acq_types,
            })

        # Tier condition
        tiers = _normalize_tier_list(tiers_raw)
        if tiers and '__ALL__' not in tiers:
            conditions.append({
                'field': 'derived_tier',
                'operator': 'in',
                'values': tiers,
            })

        # Buy category condition
        buy_cats = _normalize_buy_category_list(buy_cats_raw)
        if buy_cats and '__ALL__' not in buy_cats:
            conditions.append({
                'field': 'intake_q_buy_category',
                'operator': 'in',
                'values': buy_cats,
            })

        # Contract character condition
        characters = _normalize_character_list(character_raw)
        if characters and '__ALL__' not in characters:
            conditions.append({
                'field': 'derived_contract_character',
                'operator': 'in',
                'values': characters,
            })

        # Build the JSON rule
        if conditions:
            rule_json = json.dumps({'allOf': conditions})
        else:
            # No specific conditions — always required
            rule_json = json.dumps({'allOf': [
                {'field': 'estimated_value', 'operator': '>=', 'value': 0}
            ]})

        rule = DocumentRule(
            document_template_id=tmpl.id,
            conditions=rule_json,
            applicability=applicability if applicability in ('required', 'conditional', 'recommended') else 'required',
            priority=10,
        )
        db.session.add(rule)
        rules_created += 1

    db.session.flush()
    print(f'  Imported {templates_created} document templates with {rules_created} rules')
    return templates_created, rules_created


def import_advisory_triggers(wb):
    """Import Advisory Triggers sheet → AdvisoryTriggerRule table."""
    ws = wb['Advisory Triggers']
    count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows (3)
        trigger_id = row[0]
        if not trigger_id or not str(trigger_id).startswith('ADV-'):
            continue

        team = str(row[1]).strip() if row[1] else ''
        blocks_raw = str(row[5]).strip() if row[5] else 'No'
        blocks = 'yes' in blocks_raw.lower() or 'block' in blocks_raw.lower()

        trigger = AdvisoryTriggerRule(
            trigger_id=str(trigger_id).strip(),
            team=team,
            trigger_condition=str(row[2]).strip() if row[2] else None,
            data_needed=str(row[3]).strip() if row[3] else None,
            feeds_into_gate=str(row[4]).strip() if row[4] else None,
            blocks_gate=blocks,
            sla_days=int(row[6]) if row[6] else 5,
            escalation_to=str(row[7]).strip() if row[7] else None,
            notes=str(row[8]).strip() if row[8] else None,
        )
        db.session.add(trigger)
        count += 1

    db.session.flush()
    print(f'  Imported {count} advisory trigger rules')
    return count


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def import_all(filepath):
    """Import all rule tables from the Excel workbook.

    Clears existing rule data and re-imports from the workbook.
    Sample/transactional data (requests, CLINs, etc.) is NOT affected.

    Args:
        filepath: Path to acquisition-rules-config.xlsx

    Returns:
        dict with import counts
    """
    if not os.path.exists(filepath):
        print(f'  WARNING: Rules workbook not found at {filepath}')
        return {'error': f'File not found: {filepath}'}

    print(f'  Loading rules from {filepath}...')
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Clear existing rule tables (order matters for FK constraints)
    print('  Clearing existing rules...')
    DocumentRule.query.delete()
    DocumentTemplate.query.delete()
    ApprovalTemplateStep.query.delete()
    ApprovalTemplate.query.delete()
    IntakePath.query.delete()
    AdvisoryTriggerRule.query.delete()
    ThresholdConfig.query.delete()
    db.session.flush()

    # Import in dependency order
    counts = {}
    counts['thresholds'] = import_thresholds(wb)
    counts['intake_paths'] = import_intake_paths(wb)
    counts['approval_templates'], counts['approval_steps'] = import_approval_templates(wb)
    counts['document_templates'], counts['document_rules'] = import_document_rules(wb)
    counts['advisory_triggers'] = import_advisory_triggers(wb)

    db.session.flush()
    wb.close()

    print(f'  Rules import complete: {counts}')
    return counts
